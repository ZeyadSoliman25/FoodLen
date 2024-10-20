import openpyxl


def ModifyDataset():
    """Function to modify the dataset to make all the strings to lowercase"""

    # Load the Excel file
    file_path = 'ABBREV.xlsx'
    workbook = openpyxl.load_workbook(file_path)

    sheet = workbook['ABBREV']

    # Loop through all cells and convert text to lowercase
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):  # Only convert if the cell contains a string
                cell.value = cell.value.lower()

    # Save the modified file
    workbook.save('modified_ABBREV.xlsx')


def SearchDataset(choice):
    """Function used to search the USDA dataset to get the nutrients and calories given the choice of the user"""

    # Make the prediction string to lowercase letters
    choice = choice.lower()

    # Load the Excel file
    file_path = 'modified_ABBREV.xlsx'
    workbook = openpyxl.load_workbook(file_path)

    # Select the sheet (change 'Sheet1' to your sheet's name)
    sheet = workbook['ABBREV']

    # Define the column you want to search in (e.g., 'A' for column A)
    column_letter = 'B'

    # Loop through the column to find a match
    for cell in sheet[column_letter]:
        if choice in cell.value:
            row = cell.row  # Get the row number of the matching cell

            required_nutrients = [sheet[f'C{row}'].value, sheet[f'D{row}'].value, sheet[f'E{row}'].value,
                                  sheet[f'H{row}'].value, sheet[f'AX{row}'].value]

            return required_nutrients


def ReturnAllPossibleEntries(prediction):
    """Function used to search the USDA dataset to get all possible recipes
     and meals given the prediction of the model"""

    # Make the prediction string to lowercase letters
    prediction = prediction.lower()

    # Load the Excel file
    file_path = 'modified_ABBREV.xlsx'
    workbook = openpyxl.load_workbook(file_path)

    # Select the sheet (change 'Sheet1' to your sheet's name)
    sheet = workbook['ABBREV']

    # Define the column you want to search in (e.g., 'A' for column A)
    column_letter = 'B'

    recipes = []
    # Loop through the column to find a match
    for cell in sheet[column_letter]:
        if prediction in cell.value:
            recipes.append(cell.value)

    return recipes
